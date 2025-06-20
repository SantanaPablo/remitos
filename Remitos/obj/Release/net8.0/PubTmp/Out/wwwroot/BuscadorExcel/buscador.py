import os
import json
import xlrd  # Para .xls
import openpyxl  # Para .xlsx
import sys
from typing import List, Dict

def procesar_archivo_excel(ruta: str, archivo: str, texto_busqueda: str) -> List[Dict]:
    resultados = []
    
    try:
        if ruta.endswith('.xls'):
            libro = xlrd.open_workbook(ruta)
            hoja = libro.sheet_by_index(0)
            es_xlsx = False
        else:  # .xlsx
            libro = openpyxl.load_workbook(ruta, data_only=True)
            hoja = libro.active
            es_xlsx = True
        
        # Intentar detectar el formato del archivo
        if es_xlsx:
            # Modelo H2 (antiguo)
            if hoja.max_column > 7:
                numero_envio_cell = hoja.cell(row=2, column=8)  # H2 (fila 2, col 8)
                numero_envio = str(numero_envio_cell.value).strip() if numero_envio_cell.value else ""
            else:
                numero_envio = ""
            
            if numero_envio:  # Modelo H2
                for fila_idx in range(1, hoja.max_row + 1):
                    serial_cell = hoja.cell(row=fila_idx, column=5)  # Col E
                    usuario_cell = hoja.cell(row=fila_idx, column=6)  # Col F
                    reserva_cell = hoja.cell(row=fila_idx, column=7)  # Col G
                    
                    serial = str(serial_cell.value).lower() if serial_cell.value else ""
                    usuario = str(usuario_cell.value).lower() if usuario_cell.value else ""
                    reserva_sap = str(reserva_cell.value).lower() if reserva_cell.value else ""
                    
                    if (texto_busqueda in serial or
                        texto_busqueda in usuario or
                        texto_busqueda in reserva_sap):
                        
                        cantidad_cell = hoja.cell(row=fila_idx, column=3)  # Col C
                        cantidad = str(cantidad_cell.value) if cantidad_cell.value else ""
                        
                        resultados.append({
                            "archivo": archivo,
                            "fila": fila_idx,
                            "numero_envio": numero_envio,
                            "serial_encontrado": str(serial_cell.value) if serial_cell.value else "",
                            "usuario": str(usuario_cell.value) if usuario_cell.value else "",
                            "reserva_sap": str(reserva_cell.value) if reserva_cell.value else "",
                            "cantidad": cantidad
                        })
            else:  # Modelo Z3 (nuevo)
                numero_envio_cell = hoja.cell(row=3, column=26)  # Z3
                numero_envio = str(numero_envio_cell.value) if numero_envio_cell.value else ""
                
                # Buscar en todas las filas (desde 1)
                for fila_idx in range(1, hoja.max_row + 1):
                    serial_cell = hoja.cell(row=fila_idx, column=14)  # Col N
                    cantidad_cell = hoja.cell(row=fila_idx, column=18)  # Col R
                    usuario_cell = hoja.cell(row=fila_idx, column=19)  # Col S
                    reserva_cell = hoja.cell(row=fila_idx, column=20)  # Col T
                    
                    serial = str(serial_cell.value).lower() if serial_cell.value else ""
                    usuario = str(usuario_cell.value).lower() if usuario_cell.value else ""
                    reserva_sap = str(reserva_cell.value).lower() if reserva_cell.value else ""
                    
                    if (texto_busqueda in serial or
                        texto_busqueda in usuario or
                        texto_busqueda in reserva_sap):
                        
                        cantidad = str(cantidad_cell.value) if cantidad_cell.value else ""
                        
                        resultados.append({
                            "archivo": archivo,
                            "fila": fila_idx,
                            "numero_envio": numero_envio,
                            "serial_encontrado": str(serial_cell.value) if serial_cell.value else "",
                            "usuario": str(usuario_cell.value) if usuario_cell.value else "",
                            "reserva_sap": str(reserva_cell.value) if reserva_cell.value else "",
                            "cantidad": cantidad
                        })
        else:  # Para .xls
            # Modelo H2 (antiguo)
            numero_envio = hoja.cell_value(1, 7) if hoja.ncols > 7 else ""  # H2 = fila 1, col 7
            
            if numero_envio:  # Modelo H2
                for fila_idx in range(hoja.nrows):
                    serial = str(hoja.cell_value(fila_idx, 4)).lower()      # Col E
                    usuario = str(hoja.cell_value(fila_idx, 5)).lower()     # Col F
                    reserva_sap = str(hoja.cell_value(fila_idx, 6)).lower() # Col G

                    if (texto_busqueda in serial or
                        texto_busqueda in usuario or
                        texto_busqueda in reserva_sap):

                        cantidad = hoja.cell_value(fila_idx, 2)             # Col C

                        resultados.append({
                            "archivo": archivo,
                            "fila": fila_idx + 1,
                            "numero_envio": str(numero_envio),
                            "serial_encontrado": str(hoja.cell_value(fila_idx, 4)),
                            "usuario": str(hoja.cell_value(fila_idx, 5)),
                            "reserva_sap": str(hoja.cell_value(fila_idx, 6)),
                            "cantidad": str(cantidad)
                        })
            else:  # Modelo Z3 (nuevo)
                numero_envio = hoja.cell_value(2, 25) if hoja.ncols > 25 else ""  # Z3 = fila 2, col 25
                
                # Buscar en todas las filas
                for fila_idx in range(hoja.nrows):
                    serial = str(hoja.cell_value(fila_idx, 13)).lower()    # Col N
                    cantidad = str(hoja.cell_value(fila_idx, 17))          # Col R
                    usuario = str(hoja.cell_value(fila_idx, 18)).lower()   # Col S
                    reserva_sap = str(hoja.cell_value(fila_idx, 19)).lower() # Col T

                    if (texto_busqueda in serial or
                        texto_busqueda in usuario or
                        texto_busqueda in reserva_sap):

                        resultados.append({
                            "archivo": archivo,
                            "fila": fila_idx + 1,
                            "numero_envio": str(numero_envio),
                            "serial_encontrado": str(hoja.cell_value(fila_idx, 13)),
                            "usuario": str(hoja.cell_value(fila_idx, 18)),
                            "reserva_sap": str(hoja.cell_value(fila_idx, 19)),
                            "cantidad": cantidad
                        })
    
    except Exception as e:
        print(f"Error al procesar {archivo}: {e}", file=sys.stderr)
    
    return resultados

def buscar_texto_en_excel(directorio: str, texto_busqueda: str) -> List[Dict]:
    resultados = []
    texto_busqueda = texto_busqueda.lower()

    for archivo in os.listdir(directorio):
        if archivo.endswith(('.xls', '.xlsx')):
            ruta = os.path.join(directorio, archivo)
            resultados.extend(procesar_archivo_excel(ruta, archivo, texto_busqueda))
    
    return resultados

def imprimir_resultados_legibles(resultados: List[Dict]):
    print(f"\n {len(resultados)} resultado(s) encontrados:\n", file=sys.stderr)
    for r in resultados:
        print(
            f"{r['archivo']} - Fila {r['fila']} - Envío: {r['numero_envio']} - Usuario: {r['usuario']} - "
            f"Serial: {r['serial_encontrado']} - Reserva: {r['reserva_sap']} - Cantidad: {r['cantidad']}",
            file=sys.stderr
        )

# --- Ejecución del script ---
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps([]))
        print("Falta el parámetro de búsqueda.", file=sys.stderr)
        sys.exit(1)

    texto_a_buscar = sys.argv[1].strip().lower()
    directorio = "archivos"

    if not os.path.exists(directorio):
        print(json.dumps([]))
        print(f"No existe el directorio '{directorio}'", file=sys.stderr)
        sys.exit(1)

    resultados = buscar_texto_en_excel(directorio, texto_a_buscar)

    # Devuelve JSON al stdout
    print(json.dumps(resultados, ensure_ascii=False))

    # Muestra log legible por consola (stderr)
    imprimir_resultados_legibles(resultados)
